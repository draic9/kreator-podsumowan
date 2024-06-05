# Standard library imports
import io
import os

# Third-party library imports
import cv2
import fitz  # PyMuPDF
import numpy as np
from PIL import Image, ImageEnhance
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Interface imports
import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import filedialog

# Function to handle dropped files
def drop(event):
    files = root.tk.splitlist(event.data)
    if files:
        file_path = files[0]
        try:
            if os.path.isfile(file_path):
                label_file.config(text="Wybrano plik: " + file_path)
                global input_file_path
                input_file_path = file_path
            else:
                raise FileNotFoundError("Wybrany plik nie istnieje.")
        except Exception as e:
            label_file.config(text="Błąd: " + str(e))

# Function to handle file browsing
def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
    if file_path:
        try:
            if os.path.isfile(file_path):
                label_file.config(text="Wybrano plik: " + file_path)
                global input_file_path
                input_file_path = file_path
            else:
                raise FileNotFoundError("Wybrany plik nie istnieje.")
        except Exception as e:
            label_file.config(text="Błąd: " + str(e))

def confirm():
    try:
        if input_file_path:
            save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            if save_file_path:
                process_pdf_to_excel(input_file_path, save_file_path)
                label_file.config(text="Plik zapisano pomyślnie.")
            else:
                label_file.config(text="Operacje zapisywania anulowano")
        else:
            label_file.config(text="Nie wybrano żadnego pliku.")
    except FileNotFoundError as e:
        label_file.config(text="Błąd: " + str(e))
    except Exception as e:
        label_file.config(text="Nieobsługiwany błąd: " + str(e))

def process_pdf_to_excel(pdf_path, output_path=None):
    # Constants for crop box coordinates
    NAME_CROP_BOX = (152, 10, 325, 46)

    if output_path is None:
        output_path = os.path.splitext(pdf_path)[0] + ".xlsx"

    # Read PDF file and process images
    images = pdf_to_images(pdf_path)
    corrected_images = [correct_skew(img) for img in images]
    cropped_names = [crop_image(img, NAME_CROP_BOX) for img in corrected_images]
    enhanced_names = [enhance_image(img) for img in cropped_names]
    create_table_structure(output_path)
    insert_images_to_excel(enhanced_names, output_path)

def cancel():
    label_file.config(text="Nie wybrano jeszcze pliku.")
    global input_file_path
    input_file_path = ""

# Change label color on hover
def on_enter(event):
    label_browse.config(fg="red", font=("Helvetica", 12, "underline"))

def on_leave(event):
    label_browse.config(fg="blue", font=("Helvetica", 12))

def pdf_to_images(pdf_path):
    doc = fitz.open(pdf_path)
    images = []
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap()
        img = Image.open(io.BytesIO(pix.tobytes()))
        images.append(img)
    return images

def crop_image(image, crop_box):
    return image.crop(crop_box)

def correct_skew(image):
    # Convert PIL Image to NumPy array
    image_np = np.array(image)

    # Convert to grayscale
    gray = cv2.cvtColor(image_np, cv2.COLOR_RGB2GRAY)
    # Apply edge detection
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    # Use the Hough Line Transform to detect lines in the image
    lines = cv2.HoughLines(edges, 1, np.pi / 180, 200)
    
    # Calculate the angle of the lines
    angles = []
    if lines is not None:
        for rho, theta in lines[:, 0]:
            angle = (theta - np.pi / 2) * 180 / np.pi
            angles.append(angle)
    
    # Compute the median angle of the detected lines
    if len(angles) > 0:
        median_angle = np.median(angles)
    else:
        median_angle = 0  # If no lines are detected, assume no rotation is needed

    # Rotate the image to correct skew
    (h, w) = image_np.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, median_angle, 1.0)
    rotated = cv2.warpAffine(image_np, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    
    # Convert the corrected NumPy array back to a PIL Image
    corrected_image = Image.fromarray(cv2.cvtColor(rotated, cv2.COLOR_BGR2RGB))
    
    return corrected_image

def enhance_image(image):
    # Convert image to grayscale
    image = image.convert('L')
    
    # Enhance contrast
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(3)  # Increase contrast by a factor of 2
    
    # Enhance brightness
    enhancer = ImageEnhance.Brightness(image)
    image = enhancer.enhance(0.6)  # Decrease brightness by a factor of 0.8
    
    # Convert PIL Image to NumPy array for further processing
    image_np = np.array(image)
    
    # Normalize the image
    image_np = cv2.normalize(image_np, None, 0, 255, cv2.NORM_MINMAX)
    
    # Apply CLAHE (Contrast Limited Adaptive Histogram Equalization) to enhance the text
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    image_np = clahe.apply(image_np)
    
    # Convert back to PIL Image
    enhanced_image = Image.fromarray(image_np)
    
    return enhanced_image

def get_image_size(image):
    # Return the width and height of the image in pixels
    return image.width, image.height

def create_table_structure(output_path):
    wb = Workbook()
    ws = wb.active

    # Define a border style for the cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    # Add table header
    ws.merge_cells('A1:C1')
    header_cell = ws.cell(row=1, column=1, value="Szkoła")
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.font = Font(name='Arial', size=14, bold=True)
    header_cell.border = thin_border

    # Add border around merged cells
    for col in range(1, 4):
        ws.cell(row=1, column=col).border = thin_border

    # Add column headers
    headers = ["Numer", "Klasa", "Kwota"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.border = thin_border

    # Save the workbook
    wb.save(output_path)

def insert_images_to_excel(images, output_path):
    wb = load_workbook(output_path)
    ws = wb.active

    # Define a border style for the cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    start_row = 3
    for idx, img in enumerate(images, start=start_row):
        # Add number to the first column
        num_cell = ws.cell(row=idx, column=1, value=idx - start_row + 1)
        num_cell.border = thin_border
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.font = Font(name='Arial', size=12)
        
        # Convert the PIL image to a byte stream
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)

        # Load the image into an openpyxl Image object from the byte stream
        openpyxl_img = OpenpyxlImage(img_byte_arr)

        # Calculate cell identifier
        col_letter = get_column_letter(2)
        cell = f'{col_letter}{idx}'

        # Add the image to the spreadsheet in the second column with offsets
        openpyxl_img.anchor = f'{col_letter}{idx}'

        # Add the image to the spreadsheet in the second column
        ws.add_image(openpyxl_img, cell)

        # Adjust cell size to fit the image
        img_width, img_height = img.size
        ws.row_dimensions[idx].height = img_height * 0.8
        ws.column_dimensions['B'].width = img_width / 7.18
        
        # Apply border to the cell containing the image
        image_cell = ws.cell(row=idx, column=2)
        image_cell.border = thin_border

        # Add a cell with border to the third column (Kwota)
        kwota_cell = ws.cell(row=idx, column=3, value="")
        kwota_cell.border = thin_border
        kwota_cell.alignment = Alignment(horizontal='center', vertical='center')
        kwota_cell.font = Font(name='Arial', size=12)

    # Add the "SUMA" row
    suma_row = len(images) + start_row
    ws.merge_cells(f'A{suma_row}:B{suma_row}')
    suma_cell = ws.cell(row=suma_row, column=1, value="SUMA")
    suma_cell.alignment = Alignment(horizontal='center', vertical='center')
    suma_cell.font = Font(name='Arial', size=12, bold=True)
    suma_cell.border = thick_border

    # Add the empty cell with a thicker border to the right of "SUMA"
    total_cell = ws.cell(row=suma_row, column=3, value="")
    total_cell.border = thick_border
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_path)

# USAGE

# Create the main window
root = TkinterDnD.Tk()

# Set the title of the window
root.title("Kreator podsumowań")

# Set the background color of the window to white
root.configure(bg="white")

# Set the window size
root.geometry("400x350")

# Disable window resizing
root.resizable(False, False)
    
# Create a frame for the drag-and-drop area
frame = tk.Frame(root, bg="white smoke", bd=2, relief=tk.SUNKEN, width=350, height=200)
frame.pack(pady=(20, 10), padx=20)
frame.pack_propagate(False)  # Prevent the frame from resizing with its content

# Create a label inside the frame for instructions
label_drag = tk.Label(frame, text="Przeciągnij plik tutaj", font=("Helvetica", 14), fg="black", bg="white smoke")
label_drag.pack(pady=(40, 10))

# Create a label for "OR" with background
label_or = tk.Label(frame, text="LUB", font=("Helvetica", 12), fg="dark gray", bg="white smoke")
label_or.pack(pady=5)

# Create a horizontal line below the "OR" label
canvas_line = tk.Canvas(frame, width=300, height=2, bd=0, highlightthickness=0)
canvas_line.create_line(10, 1, 290, 1, fill="dark gray")
canvas_line.pack()

# Place the line on top of the "OR" label
canvas_line.place(in_=label_or, relx=0.5, rely=0.5, anchor="s")

# Bring the "OR" label to the front
label_or.lift()

# Create a label for "Browse file"
label_browse = tk.Label(frame, text="Wybierz z folderu", font=("Helvetica", 12), fg="blue", bg="white smoke", cursor="hand2")
label_browse.pack(pady=(10, 40))

# Create a label to display the selected file
label_file = tk.Label(root, text="Nie wybrano jeszcze pliku.", font=("Helvetica", 10), wraplength=320, fg="black", bg="white")
label_file.pack(pady=10)

# Create a frame for the buttons
button_frame = tk.Frame(root, bg="white")
button_frame.pack(fill=tk.X, padx=25, pady=10)

# Create buttons
confirm_button = tk.Button(button_frame, text="Zatwierdź", command=confirm, bg="pale green", fg="black", padx=10)
confirm_button.pack(side=tk.RIGHT)

cancel_button = tk.Button(button_frame, text="Anuluj", command=cancel, bg="lightgray", fg="black", padx=10)
cancel_button.pack(side=tk.RIGHT, padx=5)

# Enable drag and drop functionality
root.drop_target_register(DND_FILES)
root.dnd_bind('<<Drop>>', drop)

# Bind the browse label to the file dialog
label_browse.bind("<Button-1>", lambda e: browse_file())

# Bind hover events to change the color and underline
label_browse.bind("<Enter>", on_enter)
label_browse.bind("<Leave>", on_leave)

# Run the main loop
root.mainloop()